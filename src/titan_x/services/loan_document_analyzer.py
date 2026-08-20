from __future__ import annotations

import csv
import io
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from difflib import SequenceMatcher
from statistics import mean, median, pstdev
from typing import Any


def _num(value: Any) -> float:
    if value is None:
        return 0.0
    text = str(value).strip().replace(",", "").replace("₹", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    try:
        return float(text) if text else 0.0
    except ValueError:
        return 0.0


def _norm_name(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def _party_similarity(a: str, b: str) -> float:
    na, nb = _norm_name(a), _norm_name(b)
    if not na or not nb:
        return 0.0
    if na == nb or na in nb or nb in na:
        return 1.0
    return SequenceMatcher(None, na, nb).ratio()


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    return [dict(row) for row in csv.DictReader(io.StringIO(text))]


def parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(x or "").strip() for x in rows[0]]
    return [dict(zip(headers, row, strict=False)) for row in rows[1:] if any(x is not None for x in row)]


def parse_pdf_text(content: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _transaction_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for row in rows:
        lower = {str(k).strip().lower(): v for k, v in row.items()}
        dt = next((lower[k] for k in lower if k in {"date", "transaction date", "txn date", "value date"}), None)
        desc = next((lower[k] for k in lower if k in {"description", "narration", "remarks", "particulars", "transaction details"}), "")
        credit = next((lower[k] for k in lower if "credit" in k or k in {"cr", "deposit", "credit amount"}), 0)
        debit = next((lower[k] for k in lower if "debit" in k or k in {"dr", "withdrawal", "debit amount"}), 0)
        amount = next((lower[k] for k in lower if k in {"amount", "transaction amount"}), None)
        typ = str(next((lower[k] for k in lower if k in {"type", "transaction type", "cr/dr"}), "")).lower()
        amt = _num(amount)
        if amount is not None and not credit and not debit:
            if typ in {"cr", "credit", "c"}:
                credit = amt
            else:
                debit = amt
        result.append({"date": _parse_date(dt), "description": str(desc), "credit": _num(credit), "debit": _num(debit)})
    return [x for x in result if x["date"] is not None and (x["credit"] or x["debit"])]


def analyze_bank_statement(rows: list[dict[str, Any]]) -> dict[str, Any]:
    txns = _transaction_rows(rows)
    if not txns:
        return {"analyzer": "bank_statement", "status": "insufficient_data", "transactions": 0}
    credits = [x["credit"] for x in txns if x["credit"] > 0]
    debits = [x["debit"] for x in txns if x["debit"] > 0]
    monthly: dict[str, dict[str, float]] = defaultdict(lambda: {"credit": 0.0, "debit": 0.0})
    for x in txns:
        key = x["date"].strftime("%Y-%m")
        monthly[key]["credit"] += x["credit"]
        monthly[key]["debit"] += x["debit"]
    bounce_terms = ("bounce", "return", "returned", "dishonour", "dishonor", "nach return", "emi return")
    cash_terms = ("cash deposit", "cash dep", "cash withdrawal", "cash wd", "cash")
    emi_terms = ("emi", "loan repayment", "instalment", "installment")
    bounces = [x for x in txns if any(t in x["description"].lower() for t in bounce_terms)]
    cash = [x for x in txns if any(t in x["description"].lower() for t in cash_terms)]
    emi = [x for x in txns if any(t in x["description"].lower() for t in emi_terms) and x["debit"] > 0]
    month_values = list(monthly.values())
    avg_monthly_credit = mean(v["credit"] for v in month_values) if month_values else 0
    avg_monthly_debit = mean(v["debit"] for v in month_values) if month_values else 0
    cash_credit = sum(x["credit"] for x in cash)
    total_credit = sum(credits)
    cash_dependency = cash_credit / total_credit if total_credit else 0
    avg_emi = mean(x["debit"] for x in emi) if emi else 0
    return {
        "analyzer": "bank_statement",
        "status": "ok",
        "period": {"from": min(x["date"] for x in txns).isoformat(), "to": max(x["date"] for x in txns).isoformat()},
        "transactions": len(txns),
        "total_credits": round(total_credit, 2),
        "total_debits": round(sum(debits), 2),
        "avg_monthly_credits": round(avg_monthly_credit, 2),
        "avg_monthly_debits": round(avg_monthly_debit, 2),
        "median_credit": round(median(credits), 2) if credits else 0,
        "credit_volatility": round(pstdev(credits), 2) if len(credits) > 1 else 0,
        "estimated_monthly_emi": round(sum(x["debit"] for x in emi) / max(len(month_values), 1), 2),
        "cash_dependency": round(cash_dependency, 4),
        "cash_credit_total": round(cash_credit, 2),
        "bounce_count": len(bounces),
        "emi_transaction_count": len(emi),
        "monthly": {k: {a: round(b, 2) for a, b in v.items()} for k, v in sorted(monthly.items())},
        "risk_flags": (["multiple_bounces"] if len(bounces) > 1 else ["bounce_detected"] if bounces else [])
        + (["high_cash_dependency"] if cash_dependency > 0.50 else []),
        "transactions_normalized": txns,
    }


def analyze_bureau(text: str) -> dict[str, Any]:
    upper = text.upper()
    score_match = re.search(r"\b(?:CIBIL|CREDIT\s*SCORE|SCORE)\D{0,20}(\d{3})\b", upper)
    dpd = [int(x) for x in re.findall(r"\b(?:DPD|DAYS\s*PAST\s*DUE)\D{0,10}(\d{1,3})\b", upper)]
    enquiries = [int(x) for x in re.findall(r"(?:ENQUIR(?:Y|IES)|INQUIRIES)\D{0,20}(\d{1,3})\b", upper)]
    writeoff = any(term in upper for term in ("WRITE OFF", "WRITTEN OFF", "SETTLED", "SUIT FILED"))
    loan_terms = re.findall(r"(?:PERSONAL LOAN|BUSINESS LOAN|CONSUMER LOAN|HOME LOAN|GOLD LOAN|CREDIT CARD|LOAN)", upper)
    return {
        "analyzer": "bureau",
        "status": "ok" if text.strip() else "insufficient_data",
        "score": int(score_match.group(1)) if score_match else None,
        "dpd_values": dpd,
        "max_dpd": max(dpd) if dpd else 0,
        "enquiries_detected": max(enquiries) if enquiries else None,
        "loan_account_mentions": len(loan_terms),
        "adverse_flags": (["writeoff_or_settlement"] if writeoff else []) + (["dpd_detected"] if dpd and max(dpd) > 0 else []),
    }


def analyze_gst(text: str) -> dict[str, Any]:
    upper = text.upper()
    amounts = [_num(x) for x in re.findall(r"(?:TURNOVER|TAXABLE VALUE|TOTAL VALUE|INVOICE VALUE)\D{0,30}([₹,0-9]+(?:\.\d+)?)", upper)]
    gstins = re.findall(r"\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]\b", upper)
    b2b = len(re.findall(r"\bB2B\b", upper))
    b2c = len(re.findall(r"\bB2C\b", upper))
    late = len(re.findall(r"LATE|DELAYED|LATE FEE|INTEREST", upper))
    filing = len(re.findall(r"GSTR[- ]?(?:1|3B|9)\b", upper))
    return {
        "analyzer": "gst",
        "status": "ok" if text.strip() else "insufficient_data",
        "gstins": sorted(set(gstins)),
        "turnover_values_detected": [round(x, 2) for x in amounts if x > 0],
        "estimated_detected_turnover": round(sum(amounts), 2),
        "b2b_mentions": b2b,
        "b2c_mentions": b2c,
        "filing_mentions": filing,
        "late_filing_or_interest_mentions": late,
        "risk_flags": ["late_filing_signal"] if late else [],
    }


def analyze_gst_bills(text: str) -> dict[str, Any]:
    upper = text.upper()
    gstins = sorted(set(re.findall(r"\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]\b", upper)))
    invoice_numbers = sorted(set(re.findall(r"(?:INVOICE|INV)\s*(?:NO|NUMBER|#)?\s*[:.-]?\s*([A-Z0-9/-]{4,})", upper)))
    amounts = [_num(x) for x in re.findall(r"(?:TOTAL|GRAND TOTAL|INVOICE VALUE)\D{0,25}(?:₹\s*)?([0-9,]+(?:\.\d+)?)", upper)]
    parties = re.findall(r"(?:BUYER|BILL TO|CUSTOMER|PARTY|SELLER|SUPPLIER)\s*[:.-]\s*([^\n]{3,80})", upper)
    return {
        "analyzer": "gst_bill",
        "status": "ok" if text.strip() else "insufficient_data",
        "gstins": gstins,
        "invoice_numbers": invoice_numbers,
        "invoice_amounts": [round(x, 2) for x in amounts],
        "party_names": [p.strip() for p in parties],
        "duplicate_invoice_numbers": [x for x, n in Counter(invoice_numbers).items() if n > 1],
    }


def match_bank_to_gst(bank: dict[str, Any], gst_bills: dict[str, Any]) -> dict[str, Any]:
    txns = bank.get("transactions_normalized", [])
    parties = gst_bills.get("party_names", [])
    invoice_amounts = gst_bills.get("invoice_amounts", [])
    matches = []
    for txn in txns:
        if txn["credit"] <= 0:
            continue
        best_party, best_score = "", 0.0
        for party in parties:
            score = _party_similarity(txn["description"], party)
            if score > best_score:
                best_party, best_score = party, score
        amount_match = min((abs(txn["credit"] - amt) / max(amt, 1) for amt in invoice_amounts), default=999)
        if best_score >= 0.65 and amount_match <= 0.05:
            matches.append({"bank_description": txn["description"], "bank_amount": txn["credit"], "gst_party": best_party, "party_similarity": round(best_score, 3), "amount_difference_pct": round(amount_match, 4), "confidence": round(min(1.0, 0.6 * best_score + 0.4 * (1 - amount_match)), 3)})
    confidence = mean(x["confidence"] for x in matches) if matches else 0.0
    return {"analyzer": "bank_gst_match", "status": "ok", "matches": matches, "match_count": len(matches), "confidence": round(confidence, 3), "risk_flags": ["low_bank_gst_match_rate"] if txns and not matches else []}


def build_loan_analysis(bank: dict[str, Any] | None = None, bureau: dict[str, Any] | None = None, gst: dict[str, Any] | None = None, gst_bills: dict[str, Any] | None = None) -> dict[str, Any]:
    bank, bureau, gst, gst_bills = bank or {}, bureau or {}, gst or {}, gst_bills or {}
    monthly_income = float(bank.get("avg_monthly_credits", 0))
    monthly_emi = float(bank.get("estimated_monthly_emi", 0))
    foir = monthly_emi / monthly_income if monthly_income else 0.0
    score = 100
    flags = []
    if bank.get("bounce_count", 0) > 1:
        score -= 15; flags.append("multiple_bounces")
    if bank.get("cash_dependency", 0) > 0.50:
        score -= 10; flags.append("high_cash_dependency")
    if bureau.get("max_dpd", 0) > 30:
        score -= 20; flags.append("material_dpd")
    if bureau.get("adverse_flags"):
        score -= 20; flags.extend(bureau["adverse_flags"])
    if foir > 0.60:
        score -= 20; flags.append("high_foir")
    if gst_bills.get("duplicate_invoice_numbers"):
        score -= 20; flags.append("duplicate_invoice_numbers")
    score = max(0, min(100, score))
    decision = "review" if flags else "eligible" if score >= 70 else "decline"
    return {"analyzer": "loan_eligibility", "risk_score": score, "decision": decision, "monthly_income_estimate": round(monthly_income, 2), "monthly_emi_estimate": round(monthly_emi, 2), "foir": round(foir, 4), "gst_turnover_detected": gst.get("estimated_detected_turnover", 0), "risk_flags": sorted(set(flags)), "note": "Use this as an analysis signal; final lending policy must be configured and approved by the lender."}
