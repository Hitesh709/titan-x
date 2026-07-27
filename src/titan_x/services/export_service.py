import csv
import io
import math
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any

import structlog
from sqlalchemy import desc, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from titan_x.db.repository import BaseRepository
from titan_x.models.company import Company
from titan_x.models.dynamic_ai_score import DynamicAIScore
from titan_x.models.news import NewsArticle
from titan_x.models.news_nlp import NewsNLPAnalysis
from titan_x.models.paper_trading import PaperAccount, PaperPosition, SimulatedOrder
from titan_x.models.recommendation import Recommendation
from titan_x.models.watchlist import Watchlist, WatchlistItem, WatchlistMonitorEvent
from titan_x.services.paper_analytics_service import PaperAnalyticsService

logger = structlog.get_logger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import weasyprint

    HAS_WEASYPRINT = True
except ImportError:
    HAS_WEASYPRINT = False


class ExportService:
    def __init__(self, session: AsyncSession) -> None:
        self._session = session

    # ── Main entry point ──

    async def export(self, user_id: int, fmt: str) -> bytes:
        data = await self._gather_data(user_id)
        if fmt == "csv":
            return self._to_csv(data)
        elif fmt == "xlsx":
            return self._to_xlsx(data)
        elif fmt == "pdf":
            html = self._build_html(data)
            return self._to_pdf(html)
        raise ValueError(f"Unsupported format: {fmt}")

    # ── Data gathering ──

    async def _gather_data(self, user_id: int) -> dict[str, Any]:
        portfolio = await self._get_portfolio(user_id)
        performance = await self._get_performance(user_id)
        chart_svg = self._equity_curve_svg(performance.get("equity_curve", []))
        ai_explanation = await self._get_ai_explanation(user_id)
        recommendations = await self._get_recommendations()
        alerts = await self._get_alerts(user_id)
        return {
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "portfolio": portfolio,
            "performance": performance,
            "chart_svg": chart_svg,
            "ai_explanation": ai_explanation,
            "recommendations": recommendations,
            "alerts": alerts,
        }

    async def _get_portfolio(self, user_id: int) -> dict[str, Any]:
        account = (await self._session.execute(
            select(PaperAccount).where(PaperAccount.user_id == user_id)
        )).scalar_one_or_none()
        if not account:
            return {"has_account": False}
        positions = (await self._session.execute(
            select(PaperPosition).where(PaperPosition.account_id == account.id)
        )).scalars().all()
        positions_value = Decimal("0")
        pos_list = []
        for p in positions:
            mkt_val = Decimal("0")
            if p.current_price and p.quantity:
                mkt_val = p.current_price * p.quantity
                positions_value += mkt_val
            pos_list.append({
                "symbol": p.symbol,
                "quantity": p.quantity,
                "avg_price": float(p.average_price),
                "current_price": float(p.current_price) if p.current_price else None,
                "market_value": float(mkt_val),
                "unrealized_pnl": float(mkt_val - p.cost_basis) if p.current_price and p.quantity else 0,
                "realized_pnl": float(p.realized_pnl),
            })
        total_equity = account.cash_balance + positions_value
        return {
            "has_account": True,
            "cash_balance": float(account.cash_balance),
            "positions_value": float(positions_value),
            "total_equity": float(total_equity),
            "total_return": float(total_equity - account.initial_capital),
            "total_return_pct": float((total_equity - account.initial_capital) / account.initial_capital * 100) if account.initial_capital > 0 else 0.0,
            "positions": pos_list,
        }

    async def _get_performance(self, user_id: int) -> dict[str, Any]:
        svc = PaperAnalyticsService(self._session)
        analytics = await svc.compute_analytics(user_id)
        if not analytics:
            return {}
        closed = (await self._session.execute(
            select(SimulatedOrder)
            .where(SimulatedOrder.user_id == user_id, SimulatedOrder.status == "closed")
            .order_by(SimulatedOrder.exit_date)
        )).scalars().all()
        account = (await self._session.execute(
            select(PaperAccount).where(PaperAccount.user_id == user_id)
        )).scalar_one_or_none()
        if account:
            _, unrealized = await svc._get_positions_value(account.id)
            equity = svc._build_equity_curve(account.initial_capital, closed, unrealized)
            analytics["equity_curve"] = [float(v) for v in equity] if equity else []
        return analytics

    async def _get_ai_explanation(self, user_id: int) -> list[dict[str, Any]]:
        watchlisted = (await self._session.execute(
            select(WatchlistItem.symbol)
            .join(Watchlist, WatchlistItem.watchlist_id == Watchlist.id)
            .where(Watchlist.user_id == user_id)
            .distinct()
        )).scalars().all()
        results = []
        for symbol in watchlisted:
            score = (await self._session.execute(
                select(DynamicAIScore)
                .where(DynamicAIScore.symbol == symbol)
                .order_by(DynamicAIScore.as_of_date.desc())
                .limit(1)
            )).scalar_one_or_none()
            company = (await self._session.execute(
                select(Company).where(Company.symbol == symbol)
            )).scalar_one_or_none()
            if score:
                results.append({
                    "symbol": symbol,
                    "company_name": company.company_name if company else None,
                    "combined_score": score.combined_score,
                    "combined_signal": score.combined_signal,
                    "combined_confidence": score.combined_confidence,
                    "technical_score": score.technical_score,
                    "technical_signal": score.technical_signal,
                    "fundamental_score": score.fundamental_score,
                    "fundamental_signal": score.fundamental_signal,
                    "news_score": score.news_score,
                    "news_signal": score.news_signal,
                    "macro_score": score.macro_score,
                    "macro_signal": score.macro_signal,
                    "as_of_date": score.as_of_date.isoformat() if score.as_of_date else None,
                })
        return results

    async def _get_recommendations(self) -> list[dict[str, Any]]:
        rows = (await self._session.execute(
            select(Recommendation)
            .where(Recommendation.status == "active")
            .order_by(desc(Recommendation.score))
            .limit(20)
        )).scalars().all()
        return [
            {
                "symbol": r.symbol,
                "direction": r.direction,
                "confidence": r.confidence,
                "score": r.score,
                "reasoning": r.reasoning,
                "price_target": r.price_target,
                "current_price": r.current_price,
                "timeframe": r.timeframe,
                "risk_level": r.risk_level,
                "source": r.source,
                "generated_at": r.generated_at.isoformat() if r.generated_at else None,
            }
            for r in rows
        ]

    async def _get_alerts(self, user_id: int) -> list[dict[str, Any]]:
        rows = (await self._session.execute(
            select(WatchlistMonitorEvent)
            .where(WatchlistMonitorEvent.user_id == user_id)
            .order_by(desc(WatchlistMonitorEvent.triggered_at))
            .limit(20)
        )).scalars().all()
        return [
            {
                "symbol": e.symbol,
                "event_type": e.event_type,
                "severity": e.severity,
                "title": e.title,
                "message": e.message,
                "triggered_at": e.triggered_at.isoformat() if e.triggered_at else None,
            }
            for e in rows
        ]

    # ── SVG Chart ──

    def _equity_curve_svg(self, equity: list[float]) -> str:
        if len(equity) < 2:
            return "<p style='color:#a0aec0;font-size:13px;'>Not enough data for chart</p>"
        w, h = 700, 260
        pad = {"t": 20, "r": 20, "b": 30, "l": 50}
        cw, ch = w - pad["l"] - pad["r"], h - pad["t"] - pad["b"]
        mn, mx = min(equity), max(equity)
        rng = mx - mn or 1
        pts = []
        for i, v in enumerate(equity):
            x = pad["l"] + cw * i / (len(equity) - 1)
            y = pad["t"] + ch - ch * (v - mn) / rng
            pts.append(f"{x:.1f},{y:.1f}")
        polyline = " ".join(pts)
        fill_pts = f"{pts[0].split(',')[0]},{pad['t'] + ch} {polyline} {pts[-1].split(',')[0]},{pad['t'] + ch}"
        y_ticks = 4
        ylabels = []
        for i in range(y_ticks + 1):
            val = mn + rng * i / y_ticks
            y = pad["t"] + ch - ch * i / y_ticks
            ylabels.append(
                f'<text x="{pad["l"] - 8}" y="{y + 4}" text-anchor="end" font-size="10" fill="#718096">${val:,.2f}</text>'
                f'<line x1="{pad["l"]}" y1="{y}" x2="{w - pad["r"]}" y2="{y}" stroke="#e2e8f0" stroke-width="0.5"/>'
            )
        color = "#48bb78" if equity[-1] >= equity[0] else "#f56565"
        return f"""<svg width="{w}" height="{h}" viewBox="0 0 {w} {h}" style="max-width:100%;height:auto;">
{''.join(ylabels)}
<polygon points="{fill_pts}" fill="{color}15" />
<polyline points="{polyline}" fill="none" stroke="{color}" stroke-width="2" stroke-linejoin="round"/>
<circle cx="{pts[-1].split(',')[0]}" cy="{pts[-1].split(',')[1]}" r="4" fill="{color}"/>
</svg>"""

    # ── HTML Generation ──

    def _build_html(self, data: dict[str, Any]) -> str:
        p = data["portfolio"]
        perf = data["performance"]
        ai = data["ai_explanation"]
        recs = data["recommendations"]
        alerts = data["alerts"]

        pos_rows = ""
        for pos in p.get("positions", []):
            pos_rows += f"""<tr><td>{pos["symbol"]}</td><td>{pos["quantity"]}</td>
<td>${pos["avg_price"]:.2f}</td><td>${pos["current_price"]:.2f}</td>
<td>${pos["market_value"]:.2f}</td>
<td style="color:{"#48bb78" if pos["unrealized_pnl"] >= 0 else "#f56565"}">${pos["unrealized_pnl"]:.2f}</td></tr>"""

        ai_rows = ""
        for item in ai:
            ai_rows += f"""<tr><td>{item["symbol"]}</td><td>{item.get("company_name", "")}</td>
<td>{item.get("combined_signal", "")}</td><td>{item.get("combined_score", "")}</td>
<td>{item.get("combined_confidence", "")}</td>
<td>{item.get("technical_signal", "")} / {item.get("fundamental_signal", "")} / {item.get("news_signal", "")}</td></tr>"""

        rec_rows = ""
        for r in recs:
            rec_rows += f"""<tr><td>{r["symbol"]}</td><td>{r["direction"]}</td>
<td>{r.get("confidence", "")}</td><td>{r.get("score", "")}</td>
<td>{r.get("price_target", "")}</td><td>{r.get("reasoning", "")[:120]}</td></tr>"""

        alert_rows = ""
        for a in alerts:
            sev_color = {"info": "#48bb78", "warning": "#ecc94b", "critical": "#f56565"}.get(a["severity"], "#718096")
            alert_rows += f"""<tr><td>{a["symbol"]}</td><td>{a["event_type"]}</td>
<td style="color:{sev_color}">{a["severity"]}</td>
<td>{a["title"]}</td><td>{a.get("triggered_at", "")[:10]}</td></tr>"""

        return f"""<!DOCTYPE html><html lang="en"><head><meta charset="utf-8"/>
<style>
@page {{ size: A4; margin: 15mm; }}
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: 'Segoe UI', Arial, sans-serif; font-size: 12px; color: #1a202c; line-height: 1.5; }}
h1 {{ font-size: 22px; color: #2b6cb0; margin-bottom: 4px; }}
h2 {{ font-size: 16px; color: #2d3748; border-bottom: 2px solid #e2e8f0; padding-bottom: 4px; margin: 20px 0 10px; }}
h3 {{ font-size: 14px; color: #4a5568; margin: 12px 0 6px; }}
.header {{ background: linear-gradient(135deg, #2b6cb0, #2c5282); color: #fff; padding: 16px 20px; border-radius: 6px; margin-bottom: 20px; }}
.header h1 {{ color: #fff; }} .header p {{ color: #bee3f8; font-size: 12px; margin-top: 4px; }}
table {{ width: 100%; border-collapse: collapse; margin: 8px 0 14px; }}
th, td {{ border: 1px solid #e2e8f0; padding: 6px 10px; text-align: left; font-size: 11px; }}
th {{ background: #edf2f7; color: #2d3748; font-weight: 600; }}
tr:nth-child(even) {{ background: #f7fafc; }}
.card {{ background: #f7fafc; border: 1px solid #e2e8f0; border-radius: 6px; padding: 12px 16px; margin: 8px 0; }}
.card-grid {{ display: flex; gap: 10px; flex-wrap: wrap; }}
.card-grid > div {{ flex: 1; min-width: 120px; background: #fff; border: 1px solid #e2e8f0; border-radius: 4px; padding: 8px 12px; text-align: center; }}
.card-grid .label {{ font-size: 10px; color: #718096; }} .card-grid .value {{ font-size: 16px; font-weight: 700; color: #2d3748; }}
.chart-container {{ text-align: center; margin: 10px 0; }}
.footer {{ text-align: center; font-size: 10px; color: #a0aec0; margin-top: 30px; border-top: 1px solid #e2e8f0; padding-top: 10px; }}
@media print {{ .no-print {{ display: none; }} }}
</style></head><body>
<div class="header"><h1>Portfolio Export Report</h1><p>Generated: {data["exported_at"][:19]}</p></div>

<h2>Portfolio</h2>
{"<div class='card'>No paper trading account</div>" if not p.get("has_account") else f"""
<div class='card-grid'>
<div><div class='label'>Cash Balance</div><div class='value'>${p["cash_balance"]:,.2f}</div></div>
<div><div class='label'>Positions Value</div><div class='value'>${p["positions_value"]:,.2f}</div></div>
<div><div class='label'>Total Equity</div><div class='value'>${p["total_equity"]:,.2f}</div></div>
<div><div class='label'>Total Return</div><div class='value' style='color:{"#48bb78" if p["total_return"] >= 0 else "#f56565"}'>${p["total_return"]:,.2f} ({p["total_return_pct"]:+.2f}%)</div></div>
</div>
<table><thead><tr><th>Symbol</th><th>Qty</th><th>Avg Price</th><th>Current</th><th>Market Value</th><th>Unrealized P&L</th></tr></thead><tbody>{pos_rows}</tbody></table>"""}

<h2>Performance Analytics</h2>
{"<div class='card'>No performance data</div>" if not perf.get("total_trades") else f"""
<div class='card-grid'>
<div><div class='label'>Total Trades</div><div class='value'>{perf["total_trades"]}</div></div>
<div><div class='label'>Win Rate</div><div class='value'>{perf.get("win_rate", 0)*100:.1f}%</div></div>
<div><div class='label'>Profit Factor</div><div class='value'>{perf.get("profit_factor", "-")}</div></div>
<div><div class='label'>Sharpe Ratio</div><div class='value'>{perf.get("sharpe_ratio", "-")}</div></div>
<div><div class='label'>Sortino Ratio</div><div class='value'>{perf.get("sortino_ratio", "-")}</div></div>
<div><div class='label'>Max Drawdown</div><div class='value'>{perf.get("max_drawdown", "-")}</div></div>
<div><div class='label'>CAGR</div><div class='value'>{perf.get("cagr", "-")}</div></div>
</div>
<div class='chart-container'><h3>Equity Curve</h3>{data["chart_svg"]}</div>"""}

<h2>AI Explanation</h2>
{"<div class='card'>No watchlisted symbols with AI scores</div>" if not ai else f"""<table><thead><tr>
<th>Symbol</th><th>Company</th><th>Signal</th><th>Score</th><th>Confidence</th><th>Tech / Fund / News</th>
</tr></thead><tbody>{ai_rows}</tbody></table>"""}

<h2>Recommendations ({len(recs)})</h2>
{"<div class='card'>No active recommendations</div>" if not recs else f"""<table><thead><tr>
<th>Symbol</th><th>Direction</th><th>Confidence</th><th>Score</th><th>Price Target</th><th>Reasoning</th>
</tr></thead><tbody>{rec_rows}</tbody></table>"""}

<h2>Recent Alerts</h2>
{"<div class='card'>No alerts</div>" if not alerts else f"""<table><thead><tr>
<th>Symbol</th><th>Type</th><th>Severity</th><th>Title</th><th>Date</th>
</tr></thead><tbody>{alert_rows}</tbody></table>"""}

<div class='footer'>Export generated by Titan X &mdash; {data["exported_at"][:19]}</div>
</body></html>"""

    # ── Format converters ──

    def _to_csv(self, data: dict[str, Any]) -> bytes:
        output = io.StringIO()
        w = csv.writer(output)

        w.writerow(["Portfolio Export Report"])
        w.writerow(["Generated", data["exported_at"][:19]])
        w.writerow([])

        p = data["portfolio"]
        w.writerow(["Portfolio"])
        if p.get("has_account"):
            w.writerow(["Cash Balance", "Positions Value", "Total Equity", "Total Return", "Return %"])
            w.writerow([
                f"${p['cash_balance']:,.2f}", f"${p['positions_value']:,.2f}",
                f"${p['total_equity']:,.2f}", f"${p['total_return']:,.2f}",
                f"{p['total_return_pct']:+.2f}%",
            ])
            w.writerow([])
            w.writerow(["Symbol", "Qty", "Avg Price", "Current Price", "Market Value", "Unrealized P&L", "Realized P&L"])
            for pos in p.get("positions", []):
                w.writerow([
                    pos["symbol"], pos["quantity"], pos["avg_price"],
                    pos["current_price"], pos["market_value"],
                    pos["unrealized_pnl"], pos["realized_pnl"],
                ])
        else:
            w.writerow(["No paper trading account"])
        w.writerow([])

        perf = data["performance"]
        w.writerow(["Performance Analytics"])
        if perf.get("total_trades"):
            for key in ("total_trades", "winning_trades", "losing_trades", "win_rate", "profit_factor",
                        "sharpe_ratio", "sortino_ratio", "max_drawdown", "cagr", "expectancy"):
                w.writerow([key.replace("_", " ").title(), perf.get(key, "")])
        else:
            w.writerow(["No performance data"])
        w.writerow([])

        w.writerow(["AI Explanation"])
        ai = data["ai_explanation"]
        if ai:
            w.writerow(["Symbol", "Company", "Signal", "Score", "Confidence"])
            for item in ai:
                w.writerow([item["symbol"], item.get("company_name", ""), item.get("combined_signal", ""),
                           item.get("combined_score", ""), item.get("combined_confidence", "")])
        else:
            w.writerow(["No AI data"])
        w.writerow([])

        w.writerow(["Active Recommendations"])
        recs = data["recommendations"]
        if recs:
            w.writerow(["Symbol", "Direction", "Confidence", "Score", "Price Target", "Reasoning"])
            for r in recs:
                w.writerow([r["symbol"], r["direction"], r.get("confidence", ""), r.get("score", ""),
                           r.get("price_target", ""), r.get("reasoning", "")])
        else:
            w.writerow(["No recommendations"])
        w.writerow([])

        w.writerow(["Recent Alerts"])
        alerts = data["alerts"]
        if alerts:
            w.writerow(["Symbol", "Type", "Severity", "Title"])
            for a in alerts:
                w.writerow([a["symbol"], a["event_type"], a["severity"], a["title"]])
        else:
            w.writerow(["No alerts"])

        return output.getvalue().encode("utf-8-sig")

    def _to_xlsx(self, data: dict[str, Any]) -> bytes:
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is not installed")

        wb = openpyxl.Workbook()

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
        section_font = Font(bold=True, size=13, color="2B6CB0")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        def style_header(ws, row, cols):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

        # ── Sheet 1: Portfolio ──
        ws1 = wb.active
        ws1.title = "Portfolio"
        p = data["portfolio"]
        if p.get("has_account"):
            ws1.cell(row=1, column=1, value="Metric").font = section_font
            ws1.cell(row=1, column=2, value="Value").font = section_font
            metrics = [
                ("Cash Balance", f"${p['cash_balance']:,.2f}"),
                ("Positions Value", f"${p['positions_value']:,.2f}"),
                ("Total Equity", f"${p['total_equity']:,.2f}"),
                ("Total Return", f"${p['total_return']:,.2f}"),
                ("Return %", f"{p['total_return_pct']:+.2f}%"),
            ]
            for i, (k, v) in enumerate(metrics):
                ws1.cell(row=2 + i, column=1, value=k)
                ws1.cell(row=2 + i, column=2, value=v)

            pos_start = len(metrics) + 3
            ws1.cell(row=pos_start, column=1, value="Symbol")
            ws1.cell(row=pos_start, column=2, value="Qty")
            ws1.cell(row=pos_start, column=3, value="Avg Price")
            ws1.cell(row=pos_start, column=4, value="Current")
            ws1.cell(row=pos_start, column=5, value="Market Value")
            ws1.cell(row=pos_start, column=6, value="Unrealized P&L")
            style_header(ws1, pos_start, 6)
            for pi, pos in enumerate(p.get("positions", [])):
                r = pos_start + 1 + pi
                ws1.cell(row=r, column=1, value=pos["symbol"])
                ws1.cell(row=r, column=2, value=pos["quantity"])
                ws1.cell(row=r, column=3, value=pos["avg_price"])
                ws1.cell(row=r, column=4, value=pos["current_price"])
                ws1.cell(row=r, column=5, value=pos["market_value"])
                ws1.cell(row=r, column=6, value=pos["unrealized_pnl"])
        else:
            ws1.cell(row=1, column=1, value="No paper trading account")

        # ── Sheet 2: Performance ──
        ws2 = wb.create_sheet("Performance")
        perf = data["performance"]
        ws2.cell(row=1, column=1, value="Metric").font = section_font
        ws2.cell(row=1, column=2, value="Value").font = section_font
        keys = ["total_trades", "winning_trades", "losing_trades", "win_rate",
                "profit_factor", "sharpe_ratio", "sortino_ratio", "max_drawdown", "cagr", "expectancy"]
        for i, key in enumerate(keys):
            ws2.cell(row=2 + i, column=1, value=key.replace("_", " ").title())
            ws2.cell(row=2 + i, column=2, value=perf.get(key, ""))

        # ── Sheet 3: AI Explanation ──
        ws3 = wb.create_sheet("AI Explanation")
        ai = data["ai_explanation"]
        if ai:
            ws3.cell(row=1, column=1, value="Symbol")
            ws3.cell(row=1, column=2, value="Company")
            ws3.cell(row=1, column=3, value="Signal")
            ws3.cell(row=1, column=4, value="Score")
            ws3.cell(row=1, column=5, value="Confidence")
            ws3.cell(row=1, column=6, value="Technical")
            ws3.cell(row=1, column=7, value="Fundamental")
            ws3.cell(row=1, column=8, value="News")
            style_header(ws3, 1, 8)
            for i, item in enumerate(ai):
                ws3.cell(row=2 + i, column=1, value=item["symbol"])
                ws3.cell(row=2 + i, column=2, value=item.get("company_name", ""))
                ws3.cell(row=2 + i, column=3, value=item.get("combined_signal", ""))
                ws3.cell(row=2 + i, column=4, value=item.get("combined_score", ""))
                ws3.cell(row=2 + i, column=5, value=item.get("combined_confidence", ""))
                ws3.cell(row=2 + i, column=6, value=item.get("technical_signal", ""))
                ws3.cell(row=2 + i, column=7, value=item.get("fundamental_signal", ""))
                ws3.cell(row=2 + i, column=8, value=item.get("news_signal", ""))
        else:
            ws3.cell(row=1, column=1, value="No AI explanation data")

        # ── Sheet 4: Recommendations ──
        ws4 = wb.create_sheet("Recommendations")
        recs = data["recommendations"]
        if recs:
            ws4.cell(row=1, column=1, value="Symbol")
            ws4.cell(row=1, column=2, value="Direction")
            ws4.cell(row=1, column=3, value="Confidence")
            ws4.cell(row=1, column=4, value="Score")
            ws4.cell(row=1, column=5, value="Price Target")
            ws4.cell(row=1, column=6, value="Reasoning")
            style_header(ws4, 1, 6)
            for i, r in enumerate(recs):
                ws4.cell(row=2 + i, column=1, value=r["symbol"])
                ws4.cell(row=2 + i, column=2, value=r["direction"])
                ws4.cell(row=2 + i, column=3, value=r.get("confidence", ""))
                ws4.cell(row=2 + i, column=4, value=r.get("score", ""))
                ws4.cell(row=2 + i, column=5, value=r.get("price_target", ""))
                ws4.cell(row=2 + i, column=6, value=r.get("reasoning", ""))
        else:
            ws4.cell(row=1, column=1, value="No recommendations")

        # ── Sheet 5: Alerts ──
        ws5 = wb.create_sheet("Alerts")
        alerts = data["alerts"]
        if alerts:
            ws5.cell(row=1, column=1, value="Symbol")
            ws5.cell(row=1, column=2, value="Type")
            ws5.cell(row=1, column=3, value="Severity")
            ws5.cell(row=1, column=4, value="Title")
            ws5.cell(row=1, column=5, value="Date")
            style_header(ws5, 1, 5)
            for i, a in enumerate(alerts):
                ws5.cell(row=2 + i, column=1, value=a["symbol"])
                ws5.cell(row=2 + i, column=2, value=a["event_type"])
                ws5.cell(row=2 + i, column=3, value=a["severity"])
                ws5.cell(row=2 + i, column=4, value=a["title"])
                ws5.cell(row=2 + i, column=5, value=a.get("triggered_at", "")[:10])
        else:
            ws5.cell(row=1, column=1, value="No alerts")

        # Auto-width columns
        for ws in [ws1, ws2, ws3, ws4, ws5]:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _to_pdf(self, html: str) -> bytes:
        if not HAS_WEASYPRINT:
            raise RuntimeError("weasyprint is not installed")
        return weasyprint.HTML(string=html).write_pdf()
