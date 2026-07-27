import io
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal

import pytest
import pytest_asyncio
from sqlalchemy import event as sa_event, select
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine

from titan_x.db.base import Base
from titan_x.models.dynamic_ai_score import DynamicAIScore
from titan_x.models.news import NewsArticle
from titan_x.models.news_nlp import NewsNLPAnalysis
from titan_x.models.paper_trading import PaperAccount, PaperPosition, SimulatedOrder
from titan_x.models.price import DailyPrice
from titan_x.models.recommendation import Recommendation
from titan_x.models.user import User
from titan_x.models.watchlist import Watchlist, WatchlistItem, WatchlistMonitorEvent
from titan_x.services.export_service import ExportService
from titan_x.services.paper_trading_service import PaperTradingService

ASYNC_DB_URL = "sqlite+aiosqlite:///:memory:"


@pytest_asyncio.fixture
async def engine():
    eng = create_async_engine(ASYNC_DB_URL, echo=False)

    @sa_event.listens_for(eng.sync_engine, "connect")
    def set_sqlite_pragma(dbapi_connection, connection_record):
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    async with eng.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    yield eng
    await eng.dispose()


@pytest_asyncio.fixture
async def user(engine):
    async with AsyncSession(engine, expire_on_commit=False) as s:
        u = User(email="export@test.com", hashed_password="pw")
        s.add(u)
        await s.commit()
        yield u
        await s.close()


@pytest_asyncio.fixture
async def session(engine, user):
    async with AsyncSession(engine, expire_on_commit=False) as s:
        yield s


@pytest_asyncio.fixture
async def seeded_session(session, user):
    today = date.today()
    for sym, close in [("AAPL", 200), ("MSFT", 350)]:
        session.add(DailyPrice(symbol=sym, trade_date=today, open=close, high=close, low=close, close=close, volume=1_000_000))
        session.add(DailyPrice(symbol=sym, trade_date=today - timedelta(days=1), open=close - 5, high=close, low=close - 5, close=close - 5, volume=1_000_000))

    wl = Watchlist(user_id=user.id, name="My WL")
    session.add(wl)
    await session.flush()
    session.add(WatchlistItem(watchlist_id=wl.id, symbol="AAPL"))
    session.add(WatchlistItem(watchlist_id=wl.id, symbol="MSFT"))

    trading = PaperTradingService(session)
    await trading.create_account(user.id)

    session.add(DynamicAIScore(
        symbol="AAPL", as_of_date=today,
        combined_score=0.85, combined_signal="buy", combined_confidence=0.9,
        technical_score=0.8, technical_signal="buy",
        fundamental_score=0.7, fundamental_signal="buy",
        news_score=0.6, news_signal="neutral",
        macro_score=0.5, macro_signal="neutral",
    ))

    session.add(Recommendation(
        symbol="AAPL", direction="buy", confidence=0.85, score=0.9,
        reasoning="Strong fundamentals", status="active",
        price_target=250.0, current_price=200.0,
        timeframe="medium", risk_level="moderate",
    ))
    session.add(Recommendation(
        symbol="MSFT", direction="hold", confidence=0.6, score=0.65,
        reasoning="Fair valuation", status="active",
        price_target=360.0, current_price=350.0,
    ))

    wl2 = (await session.execute(select(Watchlist).where(Watchlist.name == "My WL"))).scalar_one()
    session.add(WatchlistMonitorEvent(
        user_id=user.id, watchlist_id=wl2.id, symbol="AAPL",
        event_type="news", severity="info", title="Positive news", message="AAPL news",
    ))
    session.add(WatchlistMonitorEvent(
        user_id=user.id, watchlist_id=wl2.id, symbol="MSFT",
        event_type="risk_event", severity="warning", title="Risk alert", message="MSFT risk",
    ))
    await session.commit()
    return session


@pytest.mark.asyncio
class TestDataGathering:
    async def test_gather_empty(self, session, user):
        svc = ExportService(session)
        data = await svc._gather_data(user.id)
        assert data["portfolio"]["has_account"] is False
        assert data["performance"] == {}
        assert data["ai_explanation"] == []
        assert data["recommendations"] == []
        assert data["alerts"] == []
        assert "exported_at" in data

    async def test_gather_full(self, seeded_session, user):
        svc = ExportService(seeded_session)
        data = await svc._gather_data(user.id)
        assert data["portfolio"]["has_account"] is True
        assert data["portfolio"]["cash_balance"] == 100000.0
        assert len(data["ai_explanation"]) >= 1
        assert len(data["recommendations"]) >= 2
        assert len(data["alerts"]) >= 2
        assert "chart_svg" in data


@pytest.mark.asyncio
class TestEquityCurveSVG:
    async def test_empty(self, session, user):
        svc = ExportService(session)
        svg = svc._equity_curve_svg([])
        assert "Not enough data" in svg

    async def test_single_point(self, session, user):
        svc = ExportService(session)
        svg = svc._equity_curve_svg([100000])
        assert "Not enough data" in svg

    async def test_generates_svg(self, session, user):
        svc = ExportService(session)
        svg = svc._equity_curve_svg([100000, 101000, 100500, 102000])
        assert "<svg" in svg
        assert "polyline" in svg
        assert "polygon" in svg


class TestCSV:
    def test_export_csv(self):
        svc = ExportService(None)
        data = {
            "exported_at": "2026-07-22T12:00:00",
            "portfolio": {
                "has_account": True, "cash_balance": 90000.0, "positions_value": 10000.0,
                "total_equity": 100000.0, "total_return": 0.0, "total_return_pct": 0.0,
                "positions": [
                    {"symbol": "AAPL", "quantity": 10, "avg_price": 200.0, "current_price": 210.0,
                     "market_value": 2100.0, "unrealized_pnl": 100.0, "realized_pnl": 0.0},
                ],
            },
            "performance": {"total_trades": 5, "winning_trades": 3, "losing_trades": 2,
                           "win_rate": 0.6, "profit_factor": 1.5, "sharpe_ratio": 1.2,
                           "sortino_ratio": 1.8, "max_drawdown": -0.1, "cagr": 0.15, "expectancy": 50.0},
            "ai_explanation": [{"symbol": "AAPL", "company_name": "Apple Inc", "combined_signal": "buy",
                              "combined_score": 0.85, "combined_confidence": 0.9}],
            "recommendations": [{"symbol": "AAPL", "direction": "buy", "confidence": 0.85, "score": 0.9,
                               "price_target": 250.0, "reasoning": "Strong"}],
            "alerts": [{"symbol": "AAPL", "event_type": "news", "severity": "info",
                       "title": "Positive news", "triggered_at": "2026-07-22T10:00:00"}],
            "chart_svg": "",
        }
        result = svc._to_csv(data)
        assert isinstance(result, bytes)
        text = result.decode("utf-8-sig")
        assert "Portfolio Export Report" in text
        assert "AAPL" in text
        assert "Cash Balance" in text
        assert "buy" in text

    def test_csv_no_account(self):
        svc = ExportService(None)
        data = {
            "exported_at": "2026-07-22T12:00:00",
            "portfolio": {"has_account": False},
            "performance": {},
            "ai_explanation": [],
            "recommendations": [],
            "alerts": [],
            "chart_svg": "",
        }
        result = svc._to_csv(data)
        text = result.decode("utf-8-sig")
        assert "No paper trading account" in text
        assert "No performance data" in text


class TestXLSX:
    def test_export_xlsx(self):
        svc = ExportService(None)
        data = {
            "exported_at": "2026-07-22T12:00:00",
            "portfolio": {
                "has_account": True, "cash_balance": 90000.0, "positions_value": 10000.0,
                "total_equity": 100000.0, "total_return": 0.0, "total_return_pct": 0.0,
                "positions": [
                    {"symbol": "AAPL", "quantity": 10, "avg_price": 200.0, "current_price": 210.0,
                     "market_value": 2100.0, "unrealized_pnl": 100.0, "realized_pnl": 0.0},
                ],
            },
            "performance": {"total_trades": 5, "win_rate": 0.6},
            "ai_explanation": [{"symbol": "AAPL", "company_name": "Apple Inc", "combined_signal": "buy",
                              "combined_score": 0.85, "combined_confidence": 0.9,
                              "technical_signal": "buy", "fundamental_signal": "buy", "news_signal": "neutral"}],
            "recommendations": [{"symbol": "AAPL", "direction": "buy", "confidence": 0.85, "score": 0.9,
                               "price_target": 250.0, "reasoning": "Strong"}],
            "alerts": [{"symbol": "AAPL", "event_type": "news", "severity": "info",
                       "title": "Positive news", "triggered_at": "2026-07-22T10:00:00"}],
            "chart_svg": "",
        }
        result = svc._to_xlsx(data)
        assert isinstance(result, bytes)
        assert len(result) > 0

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(result))
        sheets = wb.sheetnames
        assert "Portfolio" in sheets
        assert "Performance" in sheets
        assert "AI Explanation" in sheets
        assert "Recommendations" in sheets
        assert "Alerts" in sheets
        ws = wb["Portfolio"]
        assert ws.cell(row=1, column=1).value == "Metric"


class TestHTML:
    def test_build_html_full(self):
        svc = ExportService(None)
        data = {
            "exported_at": "2026-07-22T12:00:00",
            "portfolio": {
                "has_account": True, "cash_balance": 90000.0, "positions_value": 10000.0,
                "total_equity": 100000.0, "total_return": 5000.0, "total_return_pct": 5.0,
                "positions": [
                    {"symbol": "AAPL", "quantity": 10, "avg_price": 200.0, "current_price": 210.0,
                     "market_value": 2100.0, "unrealized_pnl": 100.0, "realized_pnl": 0.0},
                ],
            },
            "performance": {"total_trades": 5, "winning_trades": 3, "losing_trades": 2,
                           "win_rate": 0.6, "profit_factor": 1.5, "sharpe_ratio": 1.2},
            "ai_explanation": [{"symbol": "AAPL", "company_name": "Apple Inc", "combined_signal": "buy",
                              "combined_score": 0.85, "combined_confidence": 0.9,
                              "technical_signal": "buy", "fundamental_signal": "buy", "news_signal": "neutral"}],
            "recommendations": [{"symbol": "AAPL", "direction": "buy", "confidence": 0.85, "score": 0.9,
                               "price_target": 250.0, "reasoning": "Strong fundamentals"}],
            "alerts": [{"symbol": "AAPL", "event_type": "news", "severity": "info",
                       "title": "Positive news", "triggered_at": "2026-07-22T10:00:00"}],
            "chart_svg": "<svg></svg>",
        }
        html = svc._build_html(data)
        assert "<!DOCTYPE html>" in html
        assert "Portfolio Export Report" in html
        assert "AAPL" in html
        assert "Strong fundamentals" in html
        assert "Positive news" in html
        assert "<svg>" in html

    def test_build_html_empty(self):
        svc = ExportService(None)
        data = {
            "exported_at": "2026-07-22T12:00:00",
            "portfolio": {"has_account": False},
            "performance": {},
            "ai_explanation": [],
            "recommendations": [],
            "alerts": [],
            "chart_svg": "",
        }
        html = svc._build_html(data)
        assert "No paper trading account" in html
        assert "No performance data" in html
        assert "No active recommendations" in html
        assert "No AI explanation" in html or "No watchlisted symbols" in html
